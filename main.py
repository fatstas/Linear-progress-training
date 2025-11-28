import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import json
import os
from RmCalculator import RMCalculator


class TrainingProgramGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("Генератор плана тренировок - Линейная прогрессия")
        self.root.geometry("1200x750")
        self.notebook = None

        # Файл для сохранения настроек
        self.settings_file = "training_settings.json"

        # Загружаем настройки или используем значения по умолчанию
        self.load_settings()

        self.setup_ui()


        self.root.bind('<KeyPress>', self._on_key_press)

        # Фокусируем окно, чтобы оно получало события клавиатуры
        self.root.focus_set()

        # Привязываем обработчик закрытия окна
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def _on_key_press(self, event):
        """Обработка нажатия клавиш"""
        key = event.keysym.lower()

        if key == 'return':
            self.generate_plan()

    def setup_ui(self):
        # Основной фрейм
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Заголовок
        title_label = ttk.Label(main_frame, text="🏋️ Генератор плана тренировок",
                                font=("Arial", 18, "bold"))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))  # Изменил columnspan на 2

        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=1, column=0, sticky="nsew")

        # Создать основную вкладку (весь существующий интерфейс)
        main_tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(main_tab, text="📊 Программа")

        # ЛЕВАЯ КОЛОНКА - настройки
        left_frame = ttk.Frame(main_tab)
        left_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 15))

        # Фрейм ввода основных параметров (переносим в left_frame)
        input_frame = ttk.LabelFrame(left_frame, text="Основные параметры", padding="10")
        input_frame.grid(row=0, column=0, sticky="ew", pady=(0, 15))

        # Одноповторный максимум
        ttk.Label(left_frame, text="Ваш разовый максимум (кг):", font=("Arial", 11)).grid(
            row=0, column=0, sticky=tk.W, pady=8)
        self.max_weight_entry = ttk.Entry(left_frame, width=12, font=("Arial", 11))
        self.max_weight_entry.grid(row=0, column=1, sticky=tk.W, pady=8, padx=(10, 30))
        self.max_weight_entry.insert(0, "100")

        # Шаг увеличения веса
        ttk.Label(left_frame, text="Шаг увеличения веса (кг):", font=("Arial", 11)).grid(
            row=0, column=2, sticky=tk.W, pady=8)
        self.step_entry = ttk.Entry(left_frame, width=12, font=("Arial", 11))
        self.step_entry.grid(row=0, column=3, sticky=tk.W, pady=8, padx=(10, 30))
        self.step_entry.insert(0, "2.5")

        # Диапазон процентов
        range_frame = ttk.Frame(left_frame)
        range_frame.grid(row=1, column=0, columnspan=4, sticky="ew", pady=10)

        # Фрейм настройки процентовки
        self.ranges_frame = ttk.LabelFrame(left_frame, text="Настройка переходов между подходами", padding="10")
        self.ranges_frame.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(0, 15))

        # Заголовки таблицы
        headers = ["От %", "До %", "Подходы х Повторы", "Действие"]
        for col, header in enumerate(headers):
            ttk.Label(self.ranges_frame, text=header, font=("Arial", 10, "bold")).grid(
                row=0, column=col, padx=8, pady=8)

        # Переменные для хранения записей
        self.range_entries = []

        # Создаем начальные строки
        self.create_range_rows(self.ranges)

        # Кнопки управления процентовкой
        range_buttons_frame = ttk.Frame(self.ranges_frame)
        range_buttons_frame.grid(row=10, column=0, columnspan=4, pady=15)

        ttk.Button(range_buttons_frame, text="+ Добавить диапазон",
                   command=self.add_range_row).pack(side=tk.LEFT, padx=5)

        ttk.Button(range_buttons_frame, text="📚 Управление пресетами",
                   command=self.show_preset_dialog).pack(side=tk.LEFT, padx=5)

        # Кнопки генерации и экспорта
        button_frame = ttk.Frame(left_frame)
        button_frame.grid(row=3, column=0, columnspan=3, pady=20)

        ttk.Button(button_frame, text="🚀 Сгенерировать план",
                   command=self.generate_plan, style="Accent.TButton").pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="📥 Экспорт в Excel",
                   command=self.export_to_excel).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="🧹 Очистить",
                   command=self.clear_all).pack(side=tk.LEFT, padx=10)

        # Область вывода (перемещаем в правую колонку)
        output_frame = ttk.LabelFrame(main_tab, text="План тренировок", padding="10")
        output_frame.grid(row=1, column=1, sticky="nsew", pady=10)  # column=1 вместо 0

        # Текстовое поле с прокруткой
        self.output_text = tk.Text(output_frame, height=35, width=70,
                                   font=("Courier New", 9))  # Увеличил высоту и ширину
        scrollbar = ttk.Scrollbar(output_frame, orient="vertical", command=self.output_text.yview)
        self.output_text.configure(yscrollcommand=scrollbar.set)

        self.output_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))

        self.rm_calc = RMCalculator(self.notebook)
        calc_tab = self.rm_calc.create_calculator_tab()
        self.notebook.add(calc_tab, text="🧮 Калькулятор 1ПМ")

        # НАСТРОЙКА РАСПРЕДЕЛЕНИЯ ПРОСТРАНСТВА
        # Основное окно
        main_frame.columnconfigure(0, weight=1)  # Левая колонка - растягивается
        main_frame.columnconfigure(1, weight=1)  # Правая колонка - растягивается
        main_frame.rowconfigure(1, weight=1)  # Строка с контентом - растягивается

        # Левая колонка
        left_frame.columnconfigure(0, weight=1)
        left_frame.rowconfigure(0, weight=0)  # input_frame - не растягивается
        left_frame.rowconfigure(1, weight=1)  # ranges_frame - растягивается
        left_frame.rowconfigure(2, weight=0)  # button_frame - не растягивается

        # Правая колонка
        output_frame.columnconfigure(0, weight=1)
        output_frame.rowconfigure(0, weight=1)

        self.root.update_idletasks()

        # Получаем требуемую ширину и высоту содержимого
        req_width = main_frame.winfo_reqwidth() + 20  # + padding
        req_height = main_frame.winfo_reqheight() + 20  # + padding

        # Устанавливаем минимальный размер окна
        self.root.minsize(req_width, req_height)

        # Устанавливаем текущий размер окна под содержимое
        self.root.geometry(f"{req_width}x{req_height}")

    def save_current_preset(self, preset_name):
        """Сохраняет текущие настройки как пресет"""
        if not preset_name:
            return

        current_ranges = self.get_current_ranges()
        self.presets[preset_name] = {
            'ranges': current_ranges,
            'max_weight': self.max_weight_entry.get(),
            'step': self.step_entry.get()
        }
        self.save_settings()
        self.update_presets_list()

    def load_preset(self, preset_name):
        """Загружает пресет"""
        if preset_name in self.presets:
            preset = self.presets[preset_name]
            self.create_range_rows(preset['ranges'])
            self.max_weight_entry.delete(0, tk.END)
            self.max_weight_entry.insert(0, preset.get('max_weight', '100'))
            self.step_entry.delete(0, tk.END)
            self.step_entry.insert(0, preset.get('step', '2.5'))
            self.current_preset_name = preset_name

    def delete_preset(self, preset_name):
        """Удаляет пресет"""
        if preset_name in self.presets:
            del self.presets[preset_name]
            self.save_settings()
            self.update_presets_list()

    def show_preset_dialog(self):
        """Показывает диалог управления пресетами"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Управление пресетами")
        dialog.geometry("500x500")
        dialog.transient(self.root)
        dialog.grab_set()

        # Сохранение текущего пресета
        save_frame = ttk.LabelFrame(dialog, text="Сохранить текущие настройки", padding="10")
        save_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(save_frame, text="Название пресета:").pack(anchor="w")
        preset_name_entry = ttk.Entry(save_frame, width=30)
        preset_name_entry.pack(fill="x", pady=5)

        ttk.Button(save_frame, text="💾 Сохранить как новый пресет",
                   command=lambda: self.save_current_preset(preset_name_entry.get())).pack(fill="x", pady=5)

        # Список пресетов
        list_frame = ttk.LabelFrame(dialog, text="Сохраненные пресеты", padding="10")
        list_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.listbox = tk.Listbox(list_frame)
        self.listbox.pack(fill="both", expand=True, pady=5)

        for preset_name in self.presets.keys():
            self.listbox.insert(tk.END, preset_name)

        # Кнопки управления
        btn_frame = ttk.Frame(list_frame)
        btn_frame.pack(fill="x", pady=5)

        ttk.Button(btn_frame, text="📂 Загрузить выбранный",
                   command=lambda: self.load_preset(self.listbox.get(tk.ACTIVE))).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="🗑️ Удалить выбранный",
                   command=lambda: self.delete_preset(self.listbox.get(tk.ACTIVE))).pack(side="left", padx=2)
        ttk.Button(btn_frame, text="✏ Переименовать",
                   command=lambda: self.rename_preset_dialog(self.listbox.get(tk.ACTIVE))).pack(side="left", padx=2)

    def update_presets_list(self):
        """Обновляет список пресетов в listbox"""
        self.listbox.delete(0, tk.END)
        for preset_name in sorted(self.presets.keys()):
            self.listbox.insert(tk.END, preset_name)

    def rename_preset_dialog(self, old_name):
        """Диалог переименования пресета"""
        if not old_name:
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("Переименование пресета")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()

        ttk.Label(dialog, text=f"Новое название для '{old_name}':").pack(pady=10)
        new_name_entry = ttk.Entry(dialog, width=30)
        new_name_entry.pack(pady=5)
        new_name_entry.insert(0, old_name)

        def rename_preset():
            new_name = new_name_entry.get()
            if new_name and new_name != old_name:
                self.presets[new_name] = self.presets.pop(old_name)
                self.save_settings()
                self.update_presets_list()
                dialog.destroy()


        ttk.Button(dialog, text="Переименовать", command=rename_preset).pack(pady=10)
        self.update_presets_list()

    def load_settings(self):
        """Загружает настройки из файла"""
        try:
            if os.path.exists(self.settings_file):
                with open(self.settings_file, 'r', encoding='utf-8') as f:
                    settings = json.load(f)

                # Загружаем диапазоны
                self.ranges = settings.get('ranges', [
                    (50, 60, "5x10"),
                    (60, 70, "5x8"),
                    (70, 90, "5x5"),
                    (90, 100, "5x3")
                ])

                self.presets = settings.get('presets', {})
                self.current_preset_name = None

            else:
                # Настройки по умолчанию
                self.ranges = [
                    (50, 60, "5x10"),
                    (60, 70, "5x8"),
                    (70, 90, "5x5"),
                    (90, 100, "5x3")
                ]

        except Exception as e:
            print(f"Ошибка загрузки настроек: {e}")
            self.ranges = [
                (50, 60, "5x10"),
                (60, 70, "5x8"),
                (70, 90, "5x5"),
                (90, 100, "5x3")
            ]

    def save_settings(self):
        """Сохраняет настройки в файл"""
        try:
            settings = {
                'ranges': self.get_current_ranges(),
                'presets': self.presets
            }

            with open(self.settings_file, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)

        except Exception as e:
            print(f"Ошибка сохранения настроек: {e}")

    def on_closing(self):
        """Обработчик закрытия окна"""
        self.save_settings()
        self.root.destroy()


    def create_range_rows(self, ranges):
        """Создает строки для редактирования диапазонов"""
        # Очищаем существующие записи
        for widget in self.ranges_frame.grid_slaves():
            if 10 > int(widget.grid_info()["row"]) > 0:  # Все кроме заголовков
                widget.destroy()

        self.range_entries = []

        # Создаем новые строки
        for i, (min_p, max_p, reps) in enumerate(ranges, 1):
            self.add_range_row(min_p, max_p, reps, i)

    def add_range_row(self, min_p=50, max_p=60, reps="5x5", row=None):
        """Добавляет строку для редактирования диапазона"""
        if row is None:
            row = len(self.range_entries) + 1

        if len(self.range_entries) >= 9:
            return

        # Поле "От %"
        min_var = tk.StringVar(value=str(min_p))
        min_entry = ttk.Entry(self.ranges_frame, width=8, textvariable=min_var, font=("Arial", 10))
        min_entry.grid(row=row, column=0, padx=8, pady=4)

        # Поле "До %"
        max_var = tk.StringVar(value=str(max_p))
        max_entry = ttk.Entry(self.ranges_frame, width=8, textvariable=max_var, font=("Arial", 10))
        max_entry.grid(row=row, column=1, padx=8, pady=4)

        # Поле "Подходы х Повторы"
        reps_var = tk.StringVar(value=reps)
        reps_entry = ttk.Entry(self.ranges_frame, width=15, textvariable=reps_var, font=("Arial", 10))
        reps_entry.grid(row=row, column=2, padx=8, pady=4)

        # Кнопка удаления
        delete_btn = ttk.Button(self.ranges_frame, text="❌", width=3,
                                command=lambda r=row: self.delete_range_row(r))
        delete_btn.grid(row=row, column=3, padx=8, pady=4)

        self.range_entries.append({
            'row': row,
            'min_var': min_var,
            'max_var': max_var,
            'reps_var': reps_var,
            'widgets': [min_entry, max_entry, reps_entry, delete_btn]
        })

    def delete_range_row(self, row):
        """Удаляет строку диапазона"""
        # Находим запись для удаления
        entry_to_delete = None
        for entry in self.range_entries:
            if entry['row'] == row:
                entry_to_delete = entry
                break

        if entry_to_delete:
            # Удаляем виджеты
            for widget in entry_to_delete['widgets']:
                widget.destroy()
            # Удаляем из списка
            self.range_entries.remove(entry_to_delete)

            # Перенумеровываем оставшиеся строки
            for i, entry in enumerate(self.range_entries, 1):
                entry['row'] = i
                for j, widget in enumerate(entry['widgets']):
                    widget.grid(row=i, column=j, padx=8, pady=4)

    def get_current_ranges(self):
        """Возвращает текущие настройки диапазонов"""
        ranges = []
        for entry in self.range_entries:
            try:
                min_p = float(entry['min_var'].get())
                max_p = float(entry['max_var'].get())
                reps = entry['reps_var'].get()
                ranges.append((min_p, max_p, reps))
            except ValueError:
                continue
        return sorted(ranges, key=lambda x: x[0])  # Сортируем по минимальному проценту

    def apply_percent_range(self):
        """Применяет выбранный диапазон процентов"""
        try:
            start_percent = float(self.start_percent_entry.get())
            end_percent = float(self.end_percent_entry.get())

            if start_percent >= end_percent:
                messagebox.showerror("Ошибка", "Стартовый процент должен быть меньше конечного")
                return

            if start_percent < 0 or end_percent > 120:
                messagebox.showerror("Ошибка", "Проценты должны быть в диапазоне 0-120%")
                return

        except ValueError:
            messagebox.showerror("Ошибка", "Введите корректные значения процентов")
            return

    def auto_fill_ranges(self):
        """Автоматически заполняет диапазоны на основе выбранных процентов"""
        try:
            start_percent = float(self.start_percent_entry.get())
            end_percent = float(self.end_percent_entry.get())

            if start_percent >= end_percent:
                messagebox.showerror("Ошибка", "Стартовый процент должен быть меньше конечного")
                return

            # Создаем равномерные диапазоны
            range_count = 4  # Количество диапазонов
            step = (end_percent - start_percent) / range_count

            auto_ranges = []
            for i in range(range_count):
                min_p = start_percent + i * step
                max_p = start_percent + (i + 1) * step

                # Определяем подходы в зависимости от процента
                if max_p <= 60:
                    reps = "5x10"
                elif max_p <= 70:
                    reps = "5x8"
                elif max_p <= 90:
                    reps = "5x5"
                else:
                    reps = "5x3"

                auto_ranges.append((min_p, max_p, reps))

            self.create_range_rows(auto_ranges)

        except ValueError:
            messagebox.showerror("Ошибка", "Сначала установите корректный диапазон процентов")

    def reset_to_default(self):
        """Сброс к стандартным настройкам"""
        self.create_range_rows(self.ranges)

    def round_weight(self, weight):
        """Округляет вес до ближайшего кратного 2.5 кг, но не менее 20 кг"""
        rounded = round(weight / 2.5) * 2.5
        return max(rounded, 20.0)

    def find_range(self, ranges):
        min_value, max_value = 100000, 0
        for step in ranges:
            if step[0] < min_value:
                min_value = step[0]
            if step[1] > max_value:
                max_value = step[1]
        return min_value, max_value

    def generate_plan(self):
        """Генерирует план тренировок"""
        try:
            one_rep_max = float(self.max_weight_entry.get().replace(',', '.'))
            step = float(self.step_entry.get().replace(',', '.'))
            # Получаем текущие настройки диапазонов
            ranges = self.get_current_ranges()

            if not ranges:
                messagebox.showerror("Ошибка", "Добавьте хотя бы один диапазон подходов")
                return

            start_percent, end_percent = self.find_range(ranges)

            if one_rep_max <= 0:
                messagebox.showerror("Ошибка", "Вес должен быть положительным числом")
                return

            if step <= 0:
                messagebox.showerror("Ошибка", "Шаг должен быть положительным числом")
                return

            if start_percent >= end_percent:
                messagebox.showerror("Ошибка", "Стартовый процент должен быть меньше конечного")
                return



            # Определяем веса на основе процентов
            start_weight = one_rep_max * start_percent / 100
            end_weight = one_rep_max * end_percent / 100

            current_weight = self.round_weight(start_weight)
            workout_num = 1

            output = f"📈 ЛИНЕЙНАЯ ПРОГРЕССИЯ\n"
            output += f"⚡ Одноповторный максимум: {one_rep_max} кг\n"
            output += f"🎯 Диапазон: {start_percent}% - {end_percent}% от максимума\n"
            output += f"📏 Стартовый вес: {start_weight:.1f} кг\n"
            output += f"🏁 Конечный вес: {end_weight:.1f} кг\n"
            output += f"📐 Шаг увеличения: {step} кг\n"
            output += "=" * 70 + "\n"
            output += f"{'Тренировка':<12} {'Вес (кг)':<12} {'% от макс':<12} {'Подходы':<20}\n"
            output += "-" * 70 + "\n"

            plan_data = []

            while self.round_weight(current_weight) <= end_weight:
                rounded_weight = self.round_weight(current_weight)
                percentage = (rounded_weight / one_rep_max) * 100

                # Находим подходящие подходы по проценту
                sets_reps = None  # По умолчанию
                for min_p, max_p, reps in ranges:
                    if min_p < percentage <= max_p:
                        sets_reps = reps
                        break

                if rounded_weight >= one_rep_max:
                    sets_reps = ranges[-1][2]

                if rounded_weight <= start_weight:
                    sets_reps = ranges[0][2]

                if sets_reps is None:
                    break
                # Форматируем вывод
                workout_label = f"{workout_num}"
                weight_label = f"{rounded_weight:.1f}"
                percentage_label = f"{percentage:.1f}%"

                sets_label = sets_reps

                output += f"{workout_label:<12} {weight_label:<12} {percentage_label:<12} {sets_label:<20}\n"
                plan_data.append([workout_num, rounded_weight, sets_reps, percentage])

                current_weight += step
                workout_num += 1

                # Защита от бесконечного цикла
                if workout_num > 150:
                    break

            output += "=" * 70 + "\n"
            output += f"📊 Всего тренировок: {workout_num - 1}\n"
            output += f"💪 Прогресс: {start_weight:.1f}кг → {end_weight:.1f}кг\n"
            output += f"📈 Прирост: +{end_weight - start_weight:.1f}кг\n"

            self.output_text.delete(1.0, tk.END)
            self.output_text.insert(1.0, output)
            self.plan_data = plan_data
            self.current_max_weight = one_rep_max
            self.current_start_percent = start_percent
            self.current_end_percent = end_percent

        except ValueError as e:
            messagebox.showerror("Ошибка", "Введите корректные числовые значения")

    def export_to_excel(self):
        """Экспортирует план тренировок в Excel"""
        if not hasattr(self, 'plan_data'):
            messagebox.showerror("Ошибка", "Сначала сгенерируйте план тренировок")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Сохранить план тренировок"
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "План тренировок"

            # Заголовок
            ws['A1'] = "План тренировок - Линейная прогрессия"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Одноповторный максимум: {self.current_max_weight} кг"
            ws['A3'] = f"Диапазон: {self.current_start_percent}% - {self.current_end_percent}%"
            ws['A4'] = f"Шаг увеличения веса: {self.step_entry.get()} кг"

            # Данные тренировок
            start_row = 6

            headers = ["Тренировка", "Вес (кг)", "Подходы х Повторы", "% от максимума"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=start_row, column=col, value=header)

            for i, workout in enumerate(self.plan_data, start=start_row + 1):
                ws.cell(row=i, column=1, value=workout[0])  # Тренировка
                ws.cell(row=i, column=2, value=workout[1])  # Вес
                ws.cell(row=i, column=3, value=workout[2])  # Подходы
                ws.cell(row=i, column=4, value=f"{workout[3]:.1f}%")  # Процент

            # Форматирование
            for row in ws.iter_rows(min_row=start_row, max_row=len(self.plan_data) + start_row,
                                    min_col=1, max_col=4):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                         top=Side(style='thin'), bottom=Side(style='thin'))

            # Жирный шрифт для заголовков
            for cell in ws[start_row]:
                cell.font = Font(bold=True)

            # Авто-ширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(file_path)
            messagebox.showinfo("Успех", f"План тренировок сохранен в:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{str(e)}")

    def clear_all(self):
        """Очищает все поля"""
        self.max_weight_entry.delete(0, tk.END)
        self.max_weight_entry.insert(0, "100")
        self.step_entry.delete(0, tk.END)
        self.step_entry.insert(0, "2.5")
        self.start_percent_entry.delete(0, tk.END)
        self.start_percent_entry.insert(0, "50")
        self.end_percent_entry.delete(0, tk.END)
        self.end_percent_entry.insert(0, "100")
        self.output_text.delete(1.0, tk.END)
        if hasattr(self, 'plan_data'):
            del self.plan_data


def main():
    root = tk.Tk()
    app = TrainingProgramGenerator(root)
    root.mainloop()


if __name__ == "__main__":
    main()